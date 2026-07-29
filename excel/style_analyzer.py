from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border
from typing import Optional, Tuple

class StyleAnalyzer:
    """Analiza estilos de una celda para ayudar en la detección de encabezados."""

    @staticmethod
    def get_cell_style_info(cell: Cell) -> dict:
        """Extrae información de estilo relevante de una celda."""
        info = {
            'bold': False,
            'fill_color': None,
            'border_bottom': False,
            'font_size': None,
            'font_color': None,
            'is_merged': False,
        }

        # Fuente
        if cell.font:
            if isinstance(cell.font, Font):
                info['bold'] = cell.font.bold or False
                if cell.font.size:
                    info['font_size'] = cell.font.size
                if cell.font.color and hasattr(cell.font.color, 'rgb'):
                    info['font_color'] = cell.font.color.rgb

        # Relleno
        if cell.fill and isinstance(cell.fill, PatternFill):
            if cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                info['fill_color'] = cell.fill.start_color.rgb
            elif cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
                info['fill_color'] = cell.fill.fgColor.rgb

        # Borde inferior
        if cell.border and isinstance(cell.border, Border):
            if cell.border.bottom and cell.border.bottom.style:
                info['border_bottom'] = True

        return info

    @staticmethod
    def row_has_style_indicators(worksheet, row_idx: int, max_col: int) -> dict:
        """
        Escanea la fila y devuelve indicadores de estilo:
        - number_bold: cantidad de celdas en negrita
        - number_colored: cantidad de celdas con color de fondo
        - has_bottom_border: si alguna celda tiene borde inferior
        - fill_colors: conjunto de colores de fondo
        """
        result = {
            'bold_count': 0,
            'colored_count': 0,
            'has_bottom_border': False,
            'fill_colors': set(),
        }
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            style = StyleAnalyzer.get_cell_style_info(cell)
            if style['bold']:
                result['bold_count'] += 1
            if style['fill_color']:
                result['colored_count'] += 1
                result['fill_colors'].add(style['fill_color'])
            if style['border_bottom']:
                result['has_bottom_border'] = True
        return result
