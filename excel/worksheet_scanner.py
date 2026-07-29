from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Optional, Tuple
from .models import CellInfo, RowInfo, CellType
from .type_classifier import TypeClassifier
from .style_analyzer import StyleAnalyzer

class WorksheetScanner:
    """
    Escanea una hoja de Excel extrayendo información de las primeras N filas
    sin cargar toda la hoja en memoria.
    """

    def __init__(self, worksheet, max_rows: int = 100, max_cols: int = 30):
        self.worksheet = worksheet
        self.max_rows = max_rows
        self.max_cols = max_cols
        self._merged_cells = set(worksheet.merged_cells) if worksheet.merged_cells else set()

    def scan(self) -> List[RowInfo]:
        """
        Escanea las primeras max_rows filas y devuelve una lista de RowInfo.
        Solo examina hasta max_cols columnas (para rendimiento).
        """
        rows_info = []
        for row_idx in range(1, min(self.max_rows, self.worksheet.max_row) + 1):
            row_cells = []
            non_empty = 0
            text_count = 0
            number_count = 0
            date_count = 0
            empty_count = 0
            max_col = 0

            for col_idx in range(1, min(self.max_cols, self.worksheet.max_column) + 1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                is_merged = self._is_cell_merged(row_idx, col_idx)

                # Clasificar tipo
                cell_type = TypeClassifier.classify(value)

                # Crear CellInfo
                cell_info = CellInfo(
                    row=row_idx - 1,  # base 0
                    col=col_idx - 1,
                    value=str(value) if value is not None else None,
                    cell_type=cell_type,
                    is_merged=is_merged,
                )
                row_cells.append(cell_info)

                if cell_type != CellType.EMPTY:
                    non_empty += 1
                    max_col = max(col_idx, max_col)
                    if cell_type == CellType.TEXT:
                        text_count += 1
                    elif cell_type == CellType.NUMBER:
                        number_count += 1
                    elif cell_type == CellType.DATE:
                        date_count += 1
                else:
                    empty_count += 1

            # Crear RowInfo
            row_info = RowInfo(
                index=row_idx - 1,
                cells=row_cells,
                non_empty_count=non_empty,
                text_count=text_count,
                number_count=number_count,
                date_count=date_count,
                empty_count=empty_count,
                max_col=max_col,
            )
            rows_info.append(row_info)

        return rows_info

    def _is_cell_merged(self, row: int, col: int) -> bool:
        """Determina si la celda está fusionada."""
        for merged_range in self._merged_cells:
            if merged_range.min_row <= row <= merged_range.max_row and \
               merged_range.min_col <= col <= merged_range.max_col:
                return True
        return False

    def get_row_style_info(self, row_idx: int, max_col: int) -> dict:
        """Obtiene información de estilo para una fila específica."""
        return StyleAnalyzer.row_has_style_indicators(self.worksheet, row_idx, max_col)
