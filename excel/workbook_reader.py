from openpyxl import load_workbook
from typing import Generator, Optional
from pathlib import Path

class WorkbookReader:
    """
    Abre un libro Excel en modo read_only y permite iterar sobre sus hojas.
    Gestiona la liberación de recursos.
    """

    @staticmethod
    def open_workbook(filepath: Path):
        """
        Abre el libro en modo read_only y retorna el objeto workbook.
        Es responsabilidad del llamador cerrarlo.
        """
        # Forzar el uso de openpyxl para archivos .xlsx
        return load_workbook(filepath, read_only=True, data_only=True)

    @staticmethod
    def sheets(workbook):
        """Iterador sobre las hojas del libro."""
        for sheet_name in workbook.sheetnames:
            yield workbook[sheet_name], sheet_name

    @staticmethod
    def close_workbook(workbook):
        """Cierra el libro liberando recursos."""
        workbook.close()
