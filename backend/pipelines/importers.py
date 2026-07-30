"""
Módulo de Importadores - Versión profesional con detección automática de regiones 
tabulares y encabezados.
Arquitectura modular, bajo consumo de memoria, alta precisión.
🔥 MODIFICADO: Rescata el último título de la columna 0 antes del header
y lo inserta como primera fila para que el PipelineProcessor pueda heredarlo.
"""
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from dataclasses import dataclass, field
from typing import List, Optional, Dict, Any, Tuple
from enum import Enum
from pathlib import Path
import io
import tempfile
from collections import Counter
from datetime import datetime

# ============================================================
# CAPA DE DOMINIO (Modelos de datos)
# ============================================================
class CellType(Enum):
    EMPTY = "empty"
    TEXT = "text"
    NUMBER = "number"
    DATE = "date"
    BOOLEAN = "boolean"
    FORMULA = "formula"
    ERROR = "error"

@dataclass
class CellInfo:
    row: int
    col: int
    value: Optional[str] = None
    cell_type: CellType = CellType.EMPTY
    is_merged: bool = False
    bold: bool = False
    fill_color: Optional[str] = None
    border_bottom: bool = False

@dataclass
class RowInfo:
    index: int
    cells: List[CellInfo]
    non_empty_count: int = 0
    text_count: int = 0
    number_count: int = 0
    date_count: int = 0
    empty_count: int = 0
    max_col: int = 0

@dataclass
class HeaderDetectionResult:
    header_row: Optional[int]
    table_start: Optional[int]
    confidence: float
    strategy: str
    diagnostics: Dict[str, Any] = field(default_factory=dict)

@dataclass
class SheetImportResult:
    sheet_name: str
    success: bool
    header_result: Optional[HeaderDetectionResult] = None
    dataframe: Optional[pd.DataFrame] = None
    error: Optional[str] = None
    diagnostics: Dict[str, Any] = field(default_factory=dict)

@dataclass
class ImportResult:
    filename: str
    successful_sheets: List[SheetImportResult] = field(default_factory=list)
    failed_sheets: List[SheetImportResult] = field(default_factory=list)
    total_sheets: int = 0
    total_successful: int = 0
    total_failed: int = 0

# ============================================================
# TYPE CLASSIFIER
# ============================================================
class TypeClassifier:
    NUMBER_PATTERNS = [
        re.compile(r'^[+-]?\d{1,3}(?:\.\d{3})*(?:,\d+)?$'),
        re.compile(r'^[+-]?\d{1,3}(?:,\d{3})*(?:\.\d+)?$'),
        re.compile(r'^[+-]?\d+(?:[.,]\d+)?$'),
        re.compile(r'^\$?\d+(?:[.,]\d+)?$'),
        re.compile(r'^\d+(?:[.,]\d+)?%$'),
    ]
    DATE_PATTERNS = [
        re.compile(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$'),
        re.compile(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2}$'),
    ]
    BOOLEAN_TRUE = re.compile(r'^(true|yes|sí|si|1)$', re.IGNORECASE)
    BOOLEAN_FALSE = re.compile(r'^(false|no|0)$', re.IGNORECASE)
    ERROR_PATTERNS = [re.compile(r'^#\w+!?$')]

    @classmethod
    def classify(cls, value) -> CellType:
        if value is None or (isinstance(value, str) and value.strip() == ''):
            return CellType.EMPTY
        if isinstance(value, bool):
            return CellType.BOOLEAN
        if isinstance(value, (int, float)):
            return CellType.NUMBER
        if isinstance(value, datetime):
            return CellType.DATE
        if isinstance(value, str):
            stripped = value.strip()
            if stripped == '':
                return CellType.EMPTY
            for pat in cls.ERROR_PATTERNS:
                if pat.match(stripped):
                    return CellType.ERROR
            for pat in cls.DATE_PATTERNS:
                if pat.match(stripped):
                    return CellType.DATE
            for pat in cls.NUMBER_PATTERNS:
                if pat.match(stripped):
                    return CellType.NUMBER
            if cls.BOOLEAN_TRUE.match(stripped) or cls.BOOLEAN_FALSE.match(stripped):
                return CellType.BOOLEAN
            if stripped.startswith('='):
                return CellType.FORMULA
            return CellType.TEXT
        return CellType.TEXT

# ============================================================
# STYLE ANALYZER
# ============================================================
class StyleAnalyzer:
    @staticmethod
    def get_cell_style_info(cell):
        info = {
            'bold': False,
            'fill_color': None,
            'border_bottom': False,
        }
        if cell.font:
            if cell.font.bold is not None:
                info['bold'] = cell.font.bold
        if cell.fill and hasattr(cell.fill, 'start_color') and cell.fill.start_color:
            if hasattr(cell.fill.start_color, 'rgb'):
                info['fill_color'] = cell.fill.start_color.rgb
        if cell.border and cell.border.bottom and cell.border.bottom.style:
            info['border_bottom'] = True
        return info

    @staticmethod
    def row_style_info(worksheet, row_idx: int, max_col: int) -> dict:
        result = {'bold_count': 0, 'colored_count': 0, 'has_bottom_border': False}
        for col in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col)
            style = StyleAnalyzer.get_cell_style_info(cell)
            if style['bold']:
                result['bold_count'] += 1
            if style['fill_color']:
                result['colored_count'] += 1
            if style['border_bottom']:
                result['has_bottom_border'] = True
        return result

# ============================================================
# WORKSHEET SCANNER
# ============================================================
class WorksheetScanner:
    def __init__(self, worksheet, max_rows=100, max_cols=30):
        self.worksheet = worksheet
        self.max_rows = max_rows
        self.max_cols = max_cols
        self._merged_cells = set()

    def scan(self) -> List[RowInfo]:
        rows_info = []
        actual_max_col = min(self.max_cols, self.worksheet.max_column or 1)

        for row_idx in range(1, min(self.max_rows, self.worksheet.max_row) + 1):
            cells = []
            non_empty = text_count = number_count = date_count = empty_count = 0
            max_col = 0
            for col_idx in range(1, actual_max_col + 1):
                try:
                    cell = self.worksheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                except Exception:
                    value = None

                cell_type = TypeClassifier.classify(value)
                cell_info = CellInfo(
                    row=row_idx - 1,
                    col=col_idx - 1,
                    value=str(value) if value is not None else None,
                    cell_type=cell_type,
                    is_merged=False
                )
                cells.append(cell_info)
                if cell_type != CellType.EMPTY:
                    non_empty += 1
                    max_col = max(col_idx, max_col)
                    if cell_type == CellType.TEXT:
                        text_count += 1
                    elif cell_type == CellType.NUMBER:
                        number_count += 1
                    elif cell_type == CellType.DATE:
                        date_count += 1
                    else:
                        empty_count += 1

            if max_col is None:
                max_col = 0

            row_info = RowInfo(
                index=row_idx - 1,
                cells=cells,
                non_empty_count=non_empty,
                text_count=text_count,
                number_count=number_count,
                date_count=date_count,
                empty_count=empty_count,
                max_col=max_col
            )
            rows_info.append(row_info)
        return rows_info

# ============================================================
# TABLE REGION DETECTOR
# ============================================================
class TableRegionDetector:
    @staticmethod
    def detect_table_start(rows_info: List[RowInfo]) -> Optional[int]:
        if not rows_info or len(rows_info) < 2:
            return None

        valid_rows = [r for r in rows_info if r.max_col > 0 and r.non_empty_count > 0]
        if len(valid_rows) < 2:
            return None

        window_size = 5
        best_score = -1
        best_start = None

        for start_idx in range(max(0, len(valid_rows) - window_size + 1)):
            window = valid_rows[start_idx:start_idx + window_size]
            if len(window) < 2:
                continue

            max_cols_in_window = max((r.max_col or 0) for r in window)
            if max_cols_in_window == 0:
                continue

            col_ratios = {}
            for col in range(max_cols_in_window):
                types = []
                for row in window:
                    if col < len(row.cells):
                        ct = row.cells[col].cell_type
                        if ct != CellType.EMPTY:
                            types.append(ct)
                if types:
                    counter = Counter(types)
                    most_common = counter.most_common(1)[0][0]
                    ratio = counter[most_common] / len(types)
                    col_ratios[col] = (most_common, ratio)

            if col_ratios:
                avg_ratio = sum(r for _, r in col_ratios.values()) / len(col_ratios)
                non_empty_cols = sum(1 for col, (_, r) in col_ratios.items() if r > 0.5)
                score = avg_ratio * (non_empty_cols / max(1, len(window[0].cells)))
                if score > best_score:
                    best_score = score
                    best_start = start_idx

        if best_score < 0.3:
            return None
        return best_start

# ============================================================
# TAXONOMY VALIDATOR
# ============================================================
class TaxonomyValidator:
    def __init__(self):
        try:
            from backend.pipelines.detectors import ColumnMapper
            self.mapper = ColumnMapper(confidence_threshold=0.0)
        except ImportError:
            self.mapper = None

    def validate(self, headers: List[str]) -> Tuple[float, float]:
        if not headers or self.mapper is None:
            return 0.0, 0.0
        dummy = pd.DataFrame([headers], columns=headers)
        mapping = self.mapper.map_columns(dummy)
        mapped = 0
        total_conf = 0.0
        for col, (field, conf) in mapping.items():
            if field is not None:
                mapped += 1
                total_conf += conf
        total = len(mapping)
        if total == 0:
            return 0.0, 0.0
        return mapped / total, total_conf / max(1, mapped)

# ============================================================
# HEADER CANDIDATE GENERATOR
# ============================================================
class HeaderCandidateGenerator:
    @staticmethod
    def generate_candidates(rows_info: List[RowInfo], table_start: Optional[int]) -> List[int]:
        if table_start is None:
            return list(range(min(10, len(rows_info))))
        candidates = set()
        if table_start > 0:
            candidates.add(table_start - 1)
        candidates.add(table_start)
        for offset in range(-3, 4):
            row = table_start + offset
            if 0 <= row < len(rows_info):
                candidates.add(row)
        if not candidates:
            candidates.update(range(min(5, len(rows_info))))
        return sorted(candidates)

# ============================================================
# HEADER SCORER
# ============================================================
class HeaderScorer:
    def __init__(self):
        self.taxonomy_validator = TaxonomyValidator()

    def score_candidate(self, row_info: RowInfo, style_info: dict, rows_info: List[RowInfo]) -> float:
        scores = {}

        # 1. Estabilidad de tipos
        if row_info.index + 1 < len(rows_info):
            next_row = rows_info[row_info.index + 1]
            total_cols = min(row_info.max_col or 0, next_row.max_col or 0)
            if total_cols > 0:
                consistency = 0
                for col in range(total_cols):
                    if col < len(row_info.cells) and col < len(next_row.cells):
                        t1 = row_info.cells[col].cell_type
                        t2 = next_row.cells[col].cell_type
                        if t1 == t2 and t1 != CellType.EMPTY:
                            consistency += 1
                scores['type_stability'] = consistency / total_cols
            else:
                scores['type_stability'] = 0.0
        else:
            scores['type_stability'] = 0.0

        # 2. Densidad de texto vs números
        total = row_info.non_empty_count or 0
        if total > 0:
            text_ratio = row_info.text_count / total
            number_ratio = row_info.number_count / total
            scores['text_density'] = text_ratio * 0.8 - number_ratio * 0.2
        else:
            scores['text_density'] = 0.0

        # 3. Penalización por vacíos
        max_col = row_info.max_col or 0
        if max_col > 0:
            scores['empty_penalty'] = 1.0 - (row_info.empty_count / max_col)
        else:
            scores['empty_penalty'] = 0.0

        # 4. Estilos
        style_score = 0.0
        if style_info.get('bold_count', 0) > 0:
            style_score += 0.3
        if style_info.get('colored_count', 0) > 0:
            style_score += 0.3
        if style_info.get('has_bottom_border', False):
            style_score += 0.4
        scores['style'] = style_score

        # 5. Taxonomía
        if row_info.max_col > 0:
            headers = [row_info.cells[col].value for col in range(row_info.max_col) if col < len(row_info.cells)]
            if headers:
                mapeo, conf = self.taxonomy_validator.validate(headers)
                scores['taxonomy'] = mapeo * 0.7 + conf * 0.3
            else:
                scores['taxonomy'] = 0.0
        else:
            scores['taxonomy'] = 0.0

        weights = {'type_stability': 0.25, 'text_density': 0.20, 'empty_penalty': 0.15, 'style': 0.15, 'taxonomy': 0.25}
        final = 0.0
        for key, val in scores.items():
            val = max(0.0, min(1.0, val)) if isinstance(val, (int, float)) else 0.0
            final += val * weights.get(key, 0.1)
        return final

# ============================================================
# CONFIDENCE CALCULATOR
# ============================================================
class ConfidenceCalculator:
    @staticmethod
    def calculate(candidates_scores: dict, table_start: Optional[int], strategy: str = "hybrid") -> Optional[HeaderDetectionResult]:
        if not candidates_scores:
            return None
        best_row = max(candidates_scores, key=candidates_scores.get)
        best_score = candidates_scores[best_row]
        sorted_scores = sorted(candidates_scores.values(), reverse=True)
        margin = sorted_scores[0] - sorted_scores[1] if len(sorted_scores) >= 2 else 1.0
        confidence = best_score * 0.7 + margin * 0.3
        confidence = max(0.0, min(1.0, confidence))
        if best_score < 0.3:
            confidence = 0.0
        return HeaderDetectionResult(
            header_row=best_row,
            table_start=table_start,
            confidence=confidence,
            strategy=strategy,
            diagnostics={'scores': candidates_scores, 'best_score': best_score}
        )

# ============================================================
# WORKBOOK READER
# ============================================================
class WorkbookReader:
    @staticmethod
    def open_workbook(filepath):
        return load_workbook(filepath, read_only=True, data_only=True)

    @staticmethod
    def sheets(workbook):
        for sheet_name in workbook.sheetnames:
            yield workbook[sheet_name], sheet_name

    @staticmethod
    def close_workbook(workbook):
        workbook.close()

# ============================================================
# EXCEL IMPORTER (Orquestador principal)
# ============================================================
class ExcelImporter:
    def __init__(self, max_scan_rows: int = 100, max_scan_cols: int = 30):
        self.max_scan_rows = max_scan_rows
        self.max_scan_cols = max_scan_cols
        self.header_scorer = HeaderScorer()

    def import_from_bytes(self, data: bytes, filename: str) -> ImportResult:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(data)
            tmp_path = Path(tmp.name)
        try:
            return self.import_from_path(tmp_path)
        finally:
            tmp_path.unlink(missing_ok=True)

    def import_from_path(self, filepath: Path) -> ImportResult:
        result = ImportResult(filename=str(filepath))
        workbook = None
        try:
            workbook = WorkbookReader.open_workbook(filepath)
            for worksheet, sheet_name in WorkbookReader.sheets(workbook):
                try:
                    sheet_result = self._process_sheet(worksheet, sheet_name, filepath)
                    if sheet_result.success:
                        result.successful_sheets.append(sheet_result)
                    else:
                        result.failed_sheets.append(sheet_result)
                except Exception as e:
                    result.failed_sheets.append(SheetImportResult(
                        sheet_name=sheet_name,
                        success=False,
                        error=str(e)
                    ))
        except Exception as e:
            result.failed_sheets.append(SheetImportResult(
                sheet_name="<workbook>",
                success=False,
                error=f"Error al abrir el libro: {e}"
            ))
        finally:
            if workbook:
                WorkbookReader.close_workbook(workbook)
        result.total_sheets = len(result.successful_sheets) + len(result.failed_sheets)
        result.total_successful = len(result.successful_sheets)
        result.total_failed = len(result.failed_sheets)
        return result

    def _process_sheet(self, worksheet, sheet_name: str, filepath: Path) -> SheetImportResult:
        # 1. Escaneo seguro de las primeras filas
        scanner = WorksheetScanner(worksheet, self.max_scan_rows, self.max_scan_cols)
        rows_info = scanner.scan()
        if not rows_info:
            return SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                error="No se pudieron escanear filas (hoja vacía o demasiado pequeña)."
            )

        # 2. Detectar región tabular (inicio de los datos)
        table_start = TableRegionDetector.detect_table_start(rows_info)
        if table_start is None:
            return SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                error="No se pudo detectar una región tabular válida."
            )

        # 🔥 RESCATE DEL TÍTULO (último valor no vacío en columna 0 antes del header)
        # Primero necesitamos saber la fila del header. Usamos la detección.
        # Para obtenerlo, usamos el mismo enfoque que el scanner, pero necesitamos el header real.
        # Vamos a obtener la fila de header desde el scanner o la detección.
        # Como no tenemos header_row directamente, vamos a usar rows_info para inferir.
        # El header está inmediatamente antes de table_start.
        header_row_candidate = table_start
        if header_row_candidate > 0:
            # Buscar hacia arriba hasta encontrar una fila con contenido significativo
            # Simplificamos: usamos table_start - 1 como header_row
            header_row = table_start - 1
        else:
            header_row = 0

        # 🔥 Rescatar el título de la columna 0 antes del header
        last_title = None
        # Usamos el worksheet original para leer las filas antes del header
        for idx in range(header_row):
            try:
                val = worksheet.cell(row=idx + 1, column=1).value
                if pd.notna(val) and str(val).strip():
                    last_title = str(val).strip()
            except:
                pass

        # 3. Generar candidatos a encabezado (alrededor del inicio de la tabla)
        candidates = HeaderCandidateGenerator.generate_candidates(rows_info, table_start)

        # 4. Puntuar cada candidato
        scorer = HeaderScorer()
        scores = {}
        for row_idx in candidates:
            if row_idx < len(rows_info):
                row_info = rows_info[row_idx]
                style_info = StyleAnalyzer.row_style_info(worksheet, row_idx + 1, row_info.max_col or 1)
                score = scorer.score_candidate(row_info, style_info, rows_info)
                scores[row_idx] = score

        if not scores:
            return SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                error="No se generaron candidatos a encabezado."
            )

        # 5. Seleccionar el mejor candidato
        best_row = max(scores, key=scores.get)
        best_score = scores[best_row]
        confidence = min(1.0, max(0.0, best_score / 30.0))

        # 6. Si la confianza es baja, intentar fallback por taxonomía
        if confidence < 0.3:
            fallback_candidates = list(range(min(10, len(rows_info))))
            best_taxonomy_score = -1
            best_taxonomy_row = 0
            for row_idx in fallback_candidates:
                if row_idx >= len(rows_info):
                    continue
                row_info = rows_info[row_idx]
                headers = []
                for cell in row_info.cells:
                    if cell.value is not None:
                        val_str = str(cell.value).strip()
                        if val_str:
                            headers.append(val_str)
                if headers:
                    mapeo, conf = TaxonomyValidator().validate(headers)
                    taxonomy_score = mapeo * 0.7 + conf * 0.3
                    if taxonomy_score > best_taxonomy_score:
                        best_taxonomy_score = taxonomy_score
                        best_taxonomy_row = row_idx
            if best_taxonomy_score > 0.5:
                best_row = best_taxonomy_row
                confidence = max(confidence, 0.4)

        header_result = HeaderDetectionResult(
            header_row=best_row,
            table_start=table_start,
            confidence=confidence,
            strategy="hybrid",
            diagnostics={'scores': scores, 'best_score': best_score}
        )

        if confidence < 0.2:
            return SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                error="No se pudo detectar la fila de encabezados con confianza suficiente.",
                diagnostics={'header_result': header_result}
            )

        # 7. Leer la hoja con pandas a partir de la fila de encabezados detectada
        try:
            df = self._read_sheet_with_pandas(filepath, sheet_name, header_result)
        except Exception as e:
            return SheetImportResult(
                sheet_name=sheet_name,
                success=False,
                error=f"Error al leer la hoja con pandas: {e}",
                header_result=header_result
            )

        # 🔥 INYECCIÓN DEL TÍTULO HUÉRFANO COMO PRIMERA FILA
        if last_title and df is not None and not df.empty:
            # Crear una fila con el título en la primera columna y NaN en el resto
            new_row = {col: np.nan for col in df.columns}
            if len(df.columns) > 0:
                new_row[df.columns[0]] = last_title
            new_df = pd.DataFrame([new_row])
            df = pd.concat([new_df, df], ignore_index=True)

        return SheetImportResult(
            sheet_name=sheet_name,
            success=True,
            header_result=header_result,
            dataframe=df,
            diagnostics={'scores': scores}
        )

    def _read_sheet_with_pandas(self, filepath: Path, sheet_name: str, header_result: HeaderDetectionResult) -> pd.DataFrame:
        header_row = header_result.header_row
        return pd.read_excel(
            filepath,
            sheet_name=sheet_name,
            header=header_row,
            engine='openpyxl'
        )

# ============================================================
# FUNCIÓN DE ENTRADA PARA REEMPLAZAR EL IMPORTADOR ACTUAL
# ============================================================
def import_excel(data: bytes, filename: str) -> ImportResult:
    """
    Función principal para importar archivos Excel.
    Retorna un objeto ImportResult con las hojas procesadas.
    """
    importer = ExcelImporter()
    return importer.import_from_bytes(data, filename)
